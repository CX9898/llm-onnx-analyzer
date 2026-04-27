#!/usr/bin/env python3
"""
Generate multi-sheet flow statistics for multiple ONNX graphs.

Compared with `analyze_onnx_flow_stats_single.py`, this script accepts multiple ONNX
files or directories and writes all results into a single Excel workbook.

Output sheets:
- `summary`: one row per ONNX file
- one sheet per ONNX file, containing per-node details for that graph

Because CSV/TSV do not support real multi-sheet workbooks, this script outputs
an `.xlsx` file instead.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import onnx
from onnx import shape_inference
from analyze_onnx_flow_stats_single import analyze_model

try:
    from openpyxl import Workbook
except ImportError as exc:  # pragma: no cover - runtime dependency check
    raise SystemExit(
        "This script requires `openpyxl`.\n"
        "Install it with: pip install openpyxl"
    ) from exc


BASE_DETAIL_HEADERS = [
    "Stage",
    "Name",
    "Type",
]


def _safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[:\\/?*\[\]]", "_", name).strip()
    cleaned = cleaned or "sheet"
    cleaned = cleaned[:31]

    candidate = cleaned
    index = 1
    while candidate in used:
        suffix = f"_{index}"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        index += 1

    used.add(candidate)
    return candidate


def _load_model_with_shapes(model_path: Path) -> onnx.ModelProto:
    model = onnx.load(str(model_path))
    try:
        model = shape_inference.infer_shapes(model)
    except Exception:
        pass
    return model


def _extract_layer_tag(stem: str) -> str:
    parts = stem.split("_")
    if len(parts) >= 2 and parts[0] == "layer":
        return f"l{parts[1]}"
    return stem


def _extract_seq_tag(stem: str) -> str | None:
    match = re.search(r"_(\d+k)$", stem)
    return match.group(1) if match else None


def _extract_chunk_tag(stem: str) -> str | None:
    match = re.search(r"_chunk(\d+)", stem)
    return f"c{match.group(1)}" if match else None


def _suggest_sheet_name(model_path: Path) -> str:
    stem = model_path.stem
    if "ChunkGatedDeltaRule" in stem:
        return "linear_attn_ChunkRule"
    if "RecurrentGatedDeltaRule" in stem:
        return "linear_attn_RecurrentRule"
    if "_delta_net_DeltaNetChunkLayoutPrep_" in stem:
        return "delta_net_ChunkLayoutPrep"
    if "_delta_net_DeltaNetChunkScan_DeltaNetChunkStep_" in stem:
        return "delta_net_ChunkScanStep"
    if "_delta_net_DeltaNetChunkScan_" in stem:
        return "delta_net_ChunkScan"
    if "_delta_net_DeltaNetTriangularSolve_" in stem:
        return "delta_net_TriangularSolve"
    if "_delta_net_DeltaNetMaskDecay_" in stem:
        return "delta_net_MaskDecay"
    return stem


def _autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_len + 2, 10), 80)


def _write_summary_sheet(workbook: Workbook, summaries: list[dict]) -> None:
    ws = workbook.active
    if ws is None:  # pragma: no cover - Workbook() always creates one sheet
        raise RuntimeError("Workbook has no active sheet")
    ws.title = "summary"

    max_inputs = max((item["input_count"] for item in summaries), default=0)
    max_outputs = max((item["output_count"] for item in summaries), default=0)

    headers = ["Sheet", "ONNX", "Module", "NodeCount", "InputCount"]
    for idx in range(1, max_inputs + 1):
        headers.append(f"Input_{idx}")
    headers.append("OutputCount")
    for idx in range(1, max_outputs + 1):
        headers.append(f"Output_{idx}")

    ws.append(headers)

    for item in summaries:
        row = [
            item["sheet_name"],
            item["model_name"],
            _display_module_name(item["module_name"]),
            item["summary"]["node_count"],
            item["input_count"],
        ]

        for name, dtype, shape in item["inputs"]:
            row.append(f"{name}:{dtype}:{shape}")
        row.extend([""] * (max_inputs - item["input_count"]))

        row.append(item["output_count"])
        for name, dtype, shape in item["outputs"]:
            row.append(f"{name}:{dtype}:{shape}")
        row.extend([""] * (max_outputs - item["output_count"]))

        ws.append(
            row
        )

    _autosize_columns(ws)


def _split_io_field(value: str) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in value.split(" | ") if part.strip()]


def _extract_tensor_name(io_part: str) -> str:
    marker = io_part.rfind(":[")
    if marker != -1:
        return io_part[:marker].strip()
    if ":" not in io_part:
        return io_part.strip()
    return io_part.rsplit(":", 1)[0].strip()


def _extract_tensor_shape(io_part: str) -> str:
    marker = io_part.rfind(":[")
    if marker != -1:
        return io_part[marker + 1 :].strip()
    if ":" not in io_part:
        return ""
    return io_part.rsplit(":", 1)[1].strip()


def _elem_type_name(elem_type: int | None) -> str:
    if elem_type is None:
        return "?"
    pretty: dict[int, str] = {
        onnx.TensorProto.FLOAT: "float32",
        onnx.TensorProto.FLOAT16: "float16",
        onnx.TensorProto.BFLOAT16: "bfloat16",
        onnx.TensorProto.DOUBLE: "float64",
        onnx.TensorProto.INT64: "int64",
        onnx.TensorProto.INT32: "int32",
        onnx.TensorProto.INT16: "int16",
        onnx.TensorProto.INT8: "int8",
        onnx.TensorProto.UINT64: "uint64",
        onnx.TensorProto.UINT32: "uint32",
        onnx.TensorProto.UINT16: "uint16",
        onnx.TensorProto.UINT8: "uint8",
        onnx.TensorProto.BOOL: "bool",
        onnx.TensorProto.STRING: "string",
    }
    if elem_type in pretty:
        return pretty[elem_type]
    try:
        return onnx.TensorProto.DataType.Name(elem_type)
    except ValueError:
        return str(elem_type)


def _build_tensor_type_map(model_path: Path) -> dict[str, str]:
    model = _load_model_with_shapes(model_path)
    type_map: dict[str, str] = {}

    for vi in list(model.graph.input) + list(model.graph.value_info) + list(model.graph.output):
        tensor_type = vi.type.tensor_type
        elem_type = tensor_type.elem_type if tensor_type.HasField("elem_type") else None
        type_map[vi.name] = _elem_type_name(elem_type)

    for init in model.graph.initializer:
        type_map[init.name] = _elem_type_name(init.data_type)

    return type_map


def _get_graph_io_summary(model_path: Path) -> dict[str, object]:
    model = _load_model_with_shapes(model_path)
    initializer_names = {init.name for init in model.graph.initializer}

    inputs = [vi for vi in model.graph.input if vi.name not in initializer_names]
    outputs = list(model.graph.output)

    def _shape_of(value_info) -> str:
        tensor_type = value_info.type.tensor_type
        if not tensor_type.HasField("shape"):
            return "?"
        dims: list[str] = []
        for dim in tensor_type.shape.dim:
            if dim.HasField("dim_value"):
                dims.append(str(int(dim.dim_value)))
            elif dim.HasField("dim_param"):
                dims.append(dim.dim_param)
            else:
                dims.append("?")
        return "[" + ", ".join(dims) + "]"

    return {
        "input_count": len(inputs),
        "inputs": [
            (vi.name, _elem_type_name(vi.type.tensor_type.elem_type), _shape_of(vi))
            for vi in inputs
        ],
        "output_count": len(outputs),
        "outputs": [
            (vi.name, _elem_type_name(vi.type.tensor_type.elem_type), _shape_of(vi))
            for vi in outputs
        ],
    }


def _display_module_name(module_name: str) -> str:
    mapping = {
        "embedding": "TokenEmbedding",
        "rotary_embedding": "RotaryPositionEmbedding",
        "input_norm": "InputNorm",
        "delta": "LinearAttentionDeltaNet",
        "delta_main": "LinearAttentionDeltaNetMain",
        "delta_chunk_layout_prep": "DeltaNetChunkLayoutPrep",
        "delta_triangular_solve": "DeltaNetTriangularSolve",
        "delta_chunk_scan": "DeltaNetChunkScan",
        "delta_chunk_step": "DeltaNetChunkStep",
        "delta_mask_decay": "DeltaNetMaskDecay",
        "linear_attn_block": "LinearAttentionBlock",
        "chunk_gated_delta_rule": "ChunkGatedDeltaRuleSubgraph",
        "recurrent_gated_delta_rule": "RecurrentGatedDeltaRuleSubgraph",
        "post_norm": "PostMixerNorm",
        "moe_ffn": "MoEFFN",
        "moe_block": "MoEBlock",
        "self_attn": "FullAttentionSelfAttention",
        "full_attn_block": "FullAttentionBlock",
        "norm": "FinalNorm",
        "lm_head": "LMHead",
        "other": "Other",
    }
    return mapping.get(module_name, module_name)


def _classify_stage_role(
    module_name: str,
    row: dict,
    idx: int,
    total_rows: int | None = None,
) -> tuple[str, str]:
    name = str(row.get("Name", ""))
    op_type = str(row.get("Type", ""))

    if module_name == "linear_attn_block":
        if total_rows is not None and idx == total_rows - 1 and op_type == "Add":
            return "ResidualAdd", "residual_add"
        if idx <= 5:
            return "RMSNorm", "input_norm"
        delta_idx = idx - 6
        stage, role = _classify_stage_role("delta", row, delta_idx)
        if stage:
            return stage, role
        if op_type == "Add":
            return "ResidualAdd", "residual_add"
        return "LinearAttentionBlock", op_type.lower()

    if module_name == "chunk_gated_delta_rule":
        if op_type == "DeltaNetMaskDecay":
            return "MaskDecayBuild", "mask_decay"
        if op_type == "DeltaNetTriangularSolve":
            return "TriangularSolve", "triangular_solve"
        if op_type == "DeltaNetChunkStep":
            return "ChunkStepCall", "chunk_step_call"
        if op_type in {"ReduceSum", "Add", "Sqrt", "Div"}:
            return "QkNormalizationOrReduction", op_type.lower()
        return "ChunkGatedDeltaRule", op_type.lower()

    if module_name == "recurrent_gated_delta_rule":
        if op_type in {"Sqrt", "Div"}:
            return "QkNormalization", op_type.lower()
        if op_type in {"Exp", "Sub", "Mul", "ReduceSum", "ScatterND"}:
            return "RecurrentStateUpdate", op_type.lower()
        return "RecurrentGatedDeltaRule", op_type.lower()

    if module_name == "full_attn_block":
        if total_rows is not None and idx == total_rows - 1 and op_type == "Add":
            return "ResidualAdd", "residual_add"
        if idx <= 5:
            return "RMSNorm", "input_norm"
        if 6 <= idx <= 12:
            if op_type == "MatMul":
                return "RotaryPositionEmbedding", "freq_matmul"
            if op_type == "Cos":
                return "RotaryPositionEmbedding", "cos_gen"
            if op_type == "Sin":
                return "RotaryPositionEmbedding", "sin_gen"
            return "RotaryPositionEmbedding", op_type.lower()
        attn_idx = idx - 13
        stage, role = _classify_stage_role("self_attn", row, attn_idx)
        if stage:
            return stage, role
        if op_type == "Add":
            return "ResidualAdd", "residual_add"
        return "FullAttentionBlock", op_type.lower()

    if module_name == "moe_block":
        if total_rows is not None and idx == total_rows - 1 and op_type == "Add":
            return "ResidualAdd", "residual_add"
        if idx <= 5:
            return "RMSNorm", "post_norm"
        moe_idx = idx - 6
        stage, role = _classify_stage_role("moe_ffn", row, moe_idx)
        if stage:
            return stage, role
        if op_type == "Add":
            return "ResidualAdd", "residual_add"
        return "MoEBlock", op_type.lower()

    if module_name == "embedding":
        return "TokenEmbedding", "embedding_lookup"
    if module_name == "rotary_embedding":
        if op_type == "MatMul":
            return "RotaryPositionEmbedding", "freq_matmul"
        if op_type == "Cos":
            return "RotaryPositionEmbedding", "cos_gen"
        if op_type == "Sin":
            return "RotaryPositionEmbedding", "sin_gen"
        return "RotaryPositionEmbedding", op_type.lower()
    if module_name in {"input_norm", "post_norm", "norm"}:
        return "RMSNorm", "rmsnorm"
    if module_name == "lm_head":
        return "LMHeadProjection", "lm_head_proj"

    if module_name == "moe_ffn":
        if idx == 0:
            return "TokenFlatten", "flatten_tokens"
        if idx == 1:
            return "ExpertRouting", "router_logits"
        if idx == 2:
            return "ExpertRouting", "router_softmax"
        if idx == 3:
            return "ExpertRouting", "topk_select"
        if idx in {4, 5}:
            return "ExpertRouting", "topk_normalize"
        if idx == 6:
            return "ExpertRouting", "routing_onehot"
        if idx == 7:
            return "ExpertRouting", "routing_cast"
        if idx == 8:
            return "ExpertRouting", "routing_unsqueeze"
        if idx in {9, 10}:
            return "ExpertRouting", "routing_merge"
        if idx == 11:
            return "SparseExpertCompute", "experts_gate_up"
        if idx in {12, 13}:
            return "SparseExpertCompute", "experts_split_gate_up"
        if idx in {14, 15, 16}:
            return "SparseExpertCompute", "experts_activation"
        if idx in {17, 18}:
            return "SparseExpertCompute", "experts_down_or_weighted_sum"
        if idx in {19, 20, 21, 22}:
            return "SharedExpertCompute", "shared_proj"
        if idx in {23, 24, 25}:
            return "SharedExpertCompute", "shared_activation_or_gate"
        if idx in {26, 27, 28, 29}:
            return "SharedExpertCompute", "shared_down_proj"
        if idx in {30, 31, 32}:
            return "SharedGateAndMerge", "shared_gate_and_merge"
        if idx == 33:
            return "SharedGateAndMerge", "restore_shape"
        return "MoECompute", op_type.lower()

    if module_name == "self_attn":
        if name.startswith("/q_proj/"):
            return "QKVProjection", "q_proj_gate"
        if name.startswith("/k_proj/"):
            return "QKVProjection", "k_proj"
        if name.startswith("/v_proj/"):
            return "QKVProjection", "v_proj"
        if name.startswith("/q_norm/"):
            return "QKNormalization", "q_norm"
        if name.startswith("/k_norm/"):
            return "QKNormalization", "k_norm"
        if name.startswith("/o_proj/"):
            return "OutputProjection", "o_proj"
        if idx <= 23:
            return "QKVProjection", "project_and_reshape"
        if 24 <= idx <= 44:
            return "ApplyRotaryPositionEmbedding", "apply_partial_rope"
        if 45 <= idx <= 53:
            return "KVCacheUpdate", "expand_kv_cache"
        if 54 <= idx <= 57:
            return "AttentionScore", "score_and_mask"
        if idx == 58:
            return "MaskedSoftmax", "attn_softmax"
        if idx == 59:
            return "AttentionValueMix", "value_matmul"
        if 60 <= idx <= 63:
            return "AttentionOutputGate", "gate_attention_output"
        return "AttentionCore", op_type.lower()

    if module_name == "delta_main":
        if name.startswith("/input_proj_pack/"):
            return "InputProjection", "input_proj"
        if name.startswith("/causal_conv_prefill/"):
            return "CausalConvPrefill", "causal_conv"
        if name.startswith("/qkv_layout_gate_prep/"):
            return "QkvLayoutGatePrep", "qkv_layout_gate_prep"
        if op_type == "DeltaNetChunkLayoutPrep":
            return "ChunkLayoutPrep", "chunk_layout_prep"
        if op_type == "DeltaNetTriangularSolve":
            return "TriangularSolve", "triangular_solve"
        if op_type == "DeltaNetChunkScan":
            return "ChunkScan", "chunk_scan"
        if name.startswith("/gated_norm_out_proj/"):
            return "GatedNormOutProj", "gated_norm_out_proj"
        if 69 <= idx <= 79:
            return "MaskDecayInline", "mask_decay_inline"
        if 80 <= idx <= 88:
            return "TriangularPrep", "triangular_prep"
        if 89 <= idx <= 95:
            return "ChunkScanPrep", "chunk_scan_prep"
        return "DeltaNetMain", op_type.lower()

    if module_name == "delta_chunk_layout_prep":
        if op_type in {"ReduceSum", "Add", "Sqrt", "Div"}:
            return "QkNormalization", "qk_normalization"
        return "ChunkReshapeAndPack", op_type.lower()

    if module_name == "delta_mask_decay":
        if op_type == "CumSum":
            return "MaskDecayCumsum", "mask_decay_cumsum"
        if op_type == "Trilu":
            return "MaskDecayTriangularMask", "mask_decay_trilu"
        return "MaskDecayBuild", op_type.lower()

    if module_name == "delta_triangular_solve":
        if op_type in {"Constant", "Gather", "Slice", "Unsqueeze", "Squeeze", "Concat"}:
            return "TriangularSolveIndexing", "triangular_indexing"
        return "TriangularSolveCore", op_type.lower()

    if module_name == "delta_chunk_scan":
        if op_type == "DeltaNetChunkStep":
            return "ChunkStepCall", "chunk_step_call"
        return "ChunkScanIndexing", op_type.lower()

    if module_name == "delta_chunk_step":
        if op_type in {"Exp", "Sub", "Unsqueeze"}:
            return "ChunkStepDecayPrep", "chunk_step_decay_prep"
        return "ChunkStepStateUpdate", op_type.lower()

    if module_name == "delta":
        if name.startswith("/in_proj_"):
            if "qkv" in name:
                return "InputProjection", "proj_qkv"
            if "z" in name:
                return "InputProjection", "proj_z"
            if "a" in name:
                return "InputProjection", "proj_a"
            if "b" in name:
                return "InputProjection", "proj_b"
        if name.startswith("/out_proj/"):
            return "OutputProjection", "out_proj"
        if idx <= 13:
            return "ConvolutionUpdate", "conv_shift_and_mix"
        if 14 <= idx <= 29:
            return "QKVSplitReshape", "split_and_reshape_heads"
        if 30 <= idx <= 49:
            return "RecurrentParameterPrep", "compute_decay_and_normalize"
        if 50 <= idx <= 62:
            return "RecurrentStateUpdate", "state_update"
        if 63 <= idx <= 74:
            return "GatedRMSNorm", "gated_rms_norm"
        return "DeltaNetCore", op_type.lower()

    return "", ""


def _write_detail_sheet(
    workbook: Workbook,
    sheet_name: str,
    rows: list[dict],
    tensor_types: dict[str, str],
) -> None:
    max_in = max((len(_split_io_field(row["InShape"])) for row in rows), default=0)
    max_out = max((len(_split_io_field(row["OutShape"])) for row in rows), default=0)

    headers = (
        BASE_DETAIL_HEADERS
        + ["InputCount"]
        + sum(([f"InName_{idx}", f"InType_{idx}", f"InShape_{idx}"] for idx in range(1, max_in + 1)), [])
        + ["OutputCount"]
        + sum(([f"OutName_{idx}", f"OutType_{idx}", f"OutShape_{idx}"] for idx in range(1, max_out + 1)), [])
    )

    ws = workbook.create_sheet(title=sheet_name)
    ws.append(headers)

    for idx, row in enumerate(rows):
        in_parts = _split_io_field(row["InShape"])
        out_parts = _split_io_field(row["OutShape"])
        in_pairs = [
            (
                _extract_tensor_name(part),
                tensor_types.get(_extract_tensor_name(part), "?"),
                _extract_tensor_shape(part),
            )
            for part in in_parts
        ]
        out_pairs = [
            (
                _extract_tensor_name(part),
                tensor_types.get(_extract_tensor_name(part), "?"),
                _extract_tensor_shape(part),
            )
            for part in out_parts
        ]
        flat_in = [item for pair in in_pairs for item in pair]
        flat_out = [item for pair in out_pairs for item in pair]
        stage, _role = _classify_stage_role(row["Module"], row, idx, total_rows=len(rows))

        ws.append(
            [
                stage,
                row["Name"],
                row["Type"],
                len(in_pairs),
            ]
            + flat_in
            + ["", "", ""] * (max_in - len(in_pairs))
            + [len(out_pairs)]
            + flat_out
            + ["", "", ""] * (max_out - len(out_pairs))
        )

    ws.freeze_panes = "A2"
    _autosize_columns(ws)


def _expand_inputs(items: list[str]) -> list[Path]:
    paths: list[Path] = []
    for item in items:
        path = Path(item).resolve()
        if path.is_dir():
            paths.extend(sorted(path.glob("*.onnx")))
        else:
            paths.append(path)
    return paths


def _default_output_anchor(items: list[str]) -> Path:
    """优先取第一个目录形参作为默认输出根；否则取第一个文件的父目录。"""
    for item in items:
        path = Path(item).resolve()
        if path.is_dir():
            return path
    return Path(items[0]).resolve().parent


def _execution_sort_key(model_path: Path) -> tuple:
    stem = model_path.stem

    if stem.startswith("embedding"):
        return (0, 0, 0, stem)
    if stem.startswith("rotary_embedding"):
        return (1, 0, 0, stem)
    if stem.startswith("layer_"):
        parts = stem.split("_")
        try:
            layer_idx = int(parts[1])
        except (IndexError, ValueError):
            layer_idx = 10_000

        if "ChunkGatedDeltaRule" in stem:
            block_order = 2
        elif "RecurrentGatedDeltaRule" in stem:
            block_order = 2
        elif len(parts) >= 4 and parts[2] == "input" and parts[3] == "norm":
            block_order = 0
        elif re.search(r"_delta_net_\d+k$", stem):
            block_order = 1
        elif "DeltaNetChunkLayoutPrep" in stem:
            block_order = 2
        elif "DeltaNetMaskDecay" in stem:
            block_order = 3
        elif "DeltaNetTriangularSolve" in stem:
            block_order = 4
        elif "DeltaNetChunkScan_chunk" in stem and "DeltaNetChunkStep" not in stem:
            block_order = 5
        elif "DeltaNetChunkStep" in stem:
            block_order = 6
        elif len(parts) >= 5 and parts[2] == "linear" and parts[3] == "attn" and parts[4] == "block":
            block_order = 1
        elif len(parts) >= 3 and parts[2] == "delta":
            block_order = 1
        elif len(parts) >= 5 and parts[2] == "full" and parts[3] == "attn" and parts[4] == "block":
            block_order = 1
        elif len(parts) >= 4 and parts[2] == "self" and parts[3] == "attn":
            block_order = 1
        elif len(parts) >= 4 and parts[2] == "post" and parts[3] == "norm":
            block_order = 2
        elif len(parts) >= 4 and parts[2] == "moe" and parts[3] == "block":
            block_order = 3
        elif len(parts) >= 4 and parts[2] == "moe" and parts[3] == "ffn":
            block_order = 3
        else:
            block_order = 9
        return (2, layer_idx, block_order, stem)
    if stem.startswith("norm"):
        return (3, 0, 0, stem)
    if stem.startswith("lm_head"):
        return (4, 0, 0, stem)
    return (5, 0, 0, stem)


def _infer_module_name(model_path: Path) -> str:
    stem = model_path.stem

    if stem.startswith("layer_"):
        if "ChunkGatedDeltaRule" in stem:
            return "chunk_gated_delta_rule"
        if "RecurrentGatedDeltaRule" in stem:
            return "recurrent_gated_delta_rule"
        if "_delta_net_DeltaNetChunkLayoutPrep_" in stem:
            return "delta_chunk_layout_prep"
        if "_delta_net_DeltaNetChunkScan_DeltaNetChunkStep_" in stem:
            return "delta_chunk_step"
        if "_delta_net_DeltaNetChunkScan_" in stem:
            return "delta_chunk_scan"
        if "_delta_net_DeltaNetTriangularSolve_" in stem:
            return "delta_triangular_solve"
        if "_delta_net_DeltaNetMaskDecay_" in stem:
            return "delta_mask_decay"
        if re.search(r"_delta_net_\d+k$", stem):
            return "delta_main"
        parts = stem.split("_")
        if len(parts) >= 5 and parts[2] == "linear" and parts[3] == "attn" and parts[4] == "block":
            return "linear_attn_block"
        if len(parts) >= 5 and parts[2] == "full" and parts[3] == "attn" and parts[4] == "block":
            return "full_attn_block"
        if len(parts) >= 4 and parts[2] == "moe" and parts[3] == "block":
            return "moe_block"
        if len(parts) >= 4 and parts[2] == "input" and parts[3] == "norm":
            return "input_norm"
        if len(parts) >= 4 and parts[2] == "post" and parts[3] == "norm":
            return "post_norm"
        if len(parts) >= 4 and parts[2] == "self" and parts[3] == "attn":
            return "self_attn"
        if len(parts) >= 4 and parts[2] == "moe" and parts[3] == "ffn":
            return "moe_ffn"
        if len(parts) >= 3:
            return parts[2]

    if stem.startswith("rotary_embedding"):
        return "rotary_embedding"
    if stem.startswith("embedding"):
        return "embedding"
    if stem.startswith("lm_head"):
        return "lm_head"
    if stem.startswith("norm"):
        return "norm"
    return "other"


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "inputs",
        nargs="+",
        help="Paths to ONNX files or directories containing ONNX files",
    )
    parser.add_argument(
        "--out_xlsx",
        default=None,
        help=(
            "Output Excel workbook path "
            "(default: <input_dir>/onnx_flow_stats_multi.xlsx, "
            "where <input_dir> 取第一个目录形参，没有则取第一个文件的父目录)"
        ),
    )
    parser.add_argument(
        "--out_json",
        default=None,
        help="Optional JSON summary path for all models (default: not emitted)",
    )
    args = parser.parse_args()

    model_paths = _expand_inputs(args.inputs)
    if not model_paths:
        raise SystemExit("No ONNX files found in the provided inputs.")
    model_paths = sorted(model_paths, key=_execution_sort_key)

    output_anchor = _default_output_anchor(args.inputs)
    out_xlsx = (
        Path(args.out_xlsx).resolve()
        if args.out_xlsx
        else output_anchor / "onnx_flow_stats_multi.xlsx"
    )
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    used_sheet_names = {"summary"}
    summaries: list[dict] = []
    onnx_rows: list[tuple[str, list[dict], dict[str, str]]] = []

    for model_path in model_paths:
        result = analyze_model(model_path)
        io_summary = _get_graph_io_summary(model_path)
        tensor_types = _build_tensor_type_map(model_path)
        model_name = model_path.stem
        module_name = _infer_module_name(model_path)
        sheet_name = _safe_sheet_name(_suggest_sheet_name(model_path), used_sheet_names)
        summaries.append(
            {
                "sheet_name": sheet_name,
                "model_name": model_name,
                "module_name": module_name,
                "summary": result["summary"],
                "input_count": io_summary["input_count"],
                "inputs": io_summary["inputs"],
                "output_count": io_summary["output_count"],
                "outputs": io_summary["outputs"],
            }
        )

        rows_with_meta: list[dict] = []
        for row in result["rows"]:
            row_with_meta = dict(row)
            row_with_meta["ONNX"] = model_name
            row_with_meta["Module"] = module_name
            rows_with_meta.append(row_with_meta)
        onnx_rows.append((sheet_name, rows_with_meta, tensor_types))

    _write_summary_sheet(workbook, summaries)

    for sheet_name, rows, tensor_types in onnx_rows:
        _write_detail_sheet(workbook, sheet_name, rows, tensor_types)
    workbook.save(out_xlsx)

    print(f"XLSX saved : {out_xlsx}")

    if args.out_json:
        out_json = Path(args.out_json).resolve()
        out_json.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "models": summaries,
            "workbook": str(out_xlsx),
        }
        out_json.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"JSON saved : {out_json}")


if __name__ == "__main__":
    main()
